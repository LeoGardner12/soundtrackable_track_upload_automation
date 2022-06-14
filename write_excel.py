import os 
import xlsxwriter 
import openpyxl
import pprint
from pathlib import Path

  

def group_wavs():
	# print(os.listdir("all_tracks"))
	all_tracks = os.listdir("all_tracks")
	track_list_list = []
	track_list = []
	previous_track =""
	c = 0
	for track in all_tracks:
		if "-" not in track:
			track_list.append([track])
	c = 0	
	for base_track in track_list:
		for track in all_tracks:
			if base_track[0][:-4] in track and track not in track_list[c]:
				track_list[c].append(track)
		c = c+1
	return track_list

def read_isrc(filename):
	# returns a dict with the track names as keys and isrc codes and values
	xlsx_file = Path(filename)
	wb_obj = openpyxl.load_workbook(xlsx_file)
	sheet = wb_obj.active	

	row_count = sheet.max_row + 1
	current_product_name = ""
	categories = ""
	isrc_dic = {}
	for row in sheet.iter_rows(max_row = row_count):
		if row[0].value != None:
			isrc_dic[row[0].value.lower()] = row[1].value
	return isrc_dic

def read_iswc_and_tunecode(filename):
	xlsx_file = Path(filename)
	wb_obj = openpyxl.load_workbook(xlsx_file)
	sheet = wb_obj.active	

	row_count = sheet.max_row + 1
	current_product_name = ""
	categories = ""
	iswc_and_tunecode_dic = {}
	for row in sheet.iter_rows(max_row = row_count):
		if row[0].value != None:
			iswc_and_tunecode_dic[row[0].value.lower()] = [row[3].value, row[4].value]
	pprint.pprint(iswc_and_tunecode_dic)
	return iswc_and_tunecode_dic

def add_borders(letter, c, cell_format,worksheet):
	worksheet.write(letter + str(c), None,cell_format)

def write_excel(grouped_wavs):
	workbook = xlsxwriter.Workbook('test_track_information_to_add.xlsx') 

	# The workbook object is then used to add new  
	# worksheet via the add_worksheet() method. 
	worksheet = workbook.add_worksheet() 
	  
	letters = "BCDEFGHIJKLMNOPQRSTUVWXYZ"
	headings = ["Name", "Categories", "Regular price", "Type", "Download 1 name", "Choose value(s)", "ISRC Code", "PRS Tunecode", "ISWC Code", "Mins value(s)", "BPM value(s)", "Genre value(s)", "Instrument value(s)", "Mood value(s)", "Price value(s)", "Purpose value(s)", "Tempo value(s)", "Rating", "Tags", "Download 1 URL", "Artist", "Year", "Genre", "Comment", "CAE number", "Composer"]
	c = 0 
	for l in letters:
		worksheet.write(l + "1", headings[c])	
		c = c + 1

	cell_format = workbook.add_format()
	
	cell_format.set_top(1)  # This is optional when using a solid fill.
	cell_format.set_top_color("#")

	end = False
	c = 2
	for track_group in grouped_wavs:
		if len(track_group) == 1:
			for letter in letters:
				add_borders(letter,c,cell_format, worksheet)	

			worksheet.write('A' + str(c), "Product",cell_format)	
			worksheet.write('B' + str(c), track_group[0][:-4],cell_format )	
			worksheet.write('C' + str(c), "Track Only",cell_format)
			worksheet.write('D' + str(c), 10,cell_format)
			worksheet.write('E' + str(c), "Simple",cell_format)	
			worksheet.write('F' + str(c), track_group[0][:-4] + ".zip",cell_format)

			isrc_dic = read_isrc("isrc.xlsx")
			track_name = track_group[0][:-4].lower()
			try:
				worksheet.write('H' + str(c), isrc_dic[track_name],cell_format)	
			except:
				print("No ISRC code given for: " + track_name)

			iswc_and_tunecode_dic = read_iswc_and_tunecode("iswc_+_tunecode.xlsx")
			try:
				worksheet.write('I' + str(c), iswc_and_tunecode_dic[track_name][0],cell_format)	
				worksheet.write('J' + str(c), iswc_and_tunecode_dic[track_name][1],cell_format)	
			except:
				print("No ISWC or PRS Tunecode given for: " + track_group[0][:-4])


			c = c+1
			worksheet.write('A' + str(c), "Variation 1")
			c = c+1
			worksheet.write('A' + str(c), "Variation 2")

		elif len(track_group)>1:
			for letter in letters:
				add_borders(letter,c,cell_format, worksheet)	

			worksheet.write('A' + str(c), "Product",cell_format)
			worksheet.write('B' + str(c), track_group[0][:-4],cell_format)

			worksheet.write('E' + str(c), "variable",cell_format)
			worksheet.write('F' + str(c), None,cell_format)
			worksheet.write('G' + str(c), "Track Only $10, Track+Stems $15",cell_format)
			isrc_dic = read_isrc("isrc.xlsx")
			track_name = track_group[0][:-4].lower()
			try:
				worksheet.write('H' + str(c), isrc_dic[track_name],cell_format)	
			except:
				print("No ISRC code given for: " + track_name)			

			iswc_and_tunecode_dic = read_iswc_and_tunecode("iswc_+_tunecode.xlsx")
			try:
				worksheet.write('I' + str(c), iswc_and_tunecode_dic[track_name][0],cell_format)	
				worksheet.write('J' + str(c), iswc_and_tunecode_dic[track_name][1],cell_format)	
			except:
				print("No ISWC or PRS Tunecode given for: " + track_group[0][:-4])



			c +=1
			worksheet.write('A' + str(c), "Variation 1")
			worksheet.write('B' + str(c), track_group[0][:-4] + " - Track Only $10")
			worksheet.write('D' + str(c), 10)
			worksheet.write('E' + str(c), "variation, downloadable, virtual")
			worksheet.write('F' + str(c), track_group[0][:-4] + ".zip")
			worksheet.write('G' + str(c), "Track Only $10")


			c+=1
			worksheet.write('A' + str(c), "Variation 2")
			worksheet.write('B' + str(c), track_group[0][:-4] + " - Track+Stems $15" )
			worksheet.write('E' + str(c), "variation, downloadable, virtual")
			worksheet.write('D' + str(c), 15)
			worksheet.write('F' + str(c), track_group[0][:-4] + " - Stems.zip")
			worksheet.write('G' + str(c), "Track+Stems $15")



			for track in track_group:
				if "Guide" in track or "Section" in track:
					c+=1
					worksheet.write('A' + str(c), "Variation 3")
					worksheet.write('B' + str(c), track_group[0][:-4] + " - Track+Sections $15")
					worksheet.write('D' + str(c), 15)
					worksheet.write('E' + str(c), "variation, downloadable, virtual")
					worksheet.write('C' + str(c-3), "Stems+Sections Available",cell_format)
					worksheet.write('F' + str(c), track_group[0][:-4] + " - Sections.zip")
					worksheet.write('G' + str(c), "Track+Sections $15")
					worksheet.write('G' + str(c-3), "Track Only $10, Track+Stems $15, Track+Sections $15",cell_format)

					break
				else:
					worksheet.write('C' + str(c-2), "Stems Available",cell_format)

		if c == 20:
			end = True 
		c = c+1
	


	# Finally, close the Excel file 
	# via the close() method. 
	workbook.close()


grouped_wavs = group_wavs()
# # pprint.pprint(grouped_wavs)
write_excel(grouped_wavs)


