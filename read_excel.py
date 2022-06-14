from pathlib import Path
import openpyxl
import pprint
import string
import os
from mutagen.mp3 import MP3
from decimal import Decimal
import time
import wave
import contextlib



def read_excel(excel_filename, src_track_directory):
	xlsx_file = Path(excel_filename)
	wb_obj = openpyxl.load_workbook(xlsx_file)
	sheet = wb_obj.active	
	track_list = []

	row_count = sheet.max_row + 1
	current_product_name = ""
	categories = ""
	for row in sheet.iter_rows(max_row = row_count):

		is_product_or_variation =""
		if row[0].value != None:
			is_product_or_variation = row[0].value.lower()

		# if the row is a variation but the product has no variations continue to next row
		if "variation" in is_product_or_variation and categories == "Track Only":
			continue

		row_list = []
		track_info = {}
		#create dic keys from excel headings 
		for heading in sheet['1']:
			if heading.value == None:
				continue
			track_info[heading.value.lower().replace(" ", "_")] = ""

		c = 1
		for key in track_info.keys():	
			track_info[key] = row[c].value
			if row[0].value != None:
				if key == "name" and "Variation 1" in row[0].value:
					track_info[key] = current_product_name + " - Track Only $10"
				if key == "name" and "Variation 2" in row[0].value:
					track_info[key] = current_product_name + " - Track+Stems $15"
				if key == "name" and "Variation 3" in row[0].value:
					track_info[key] = current_product_name + " - Track+Sections $15"
				if key == "attribute_1_value(s)" and "product" in row[0].value.lower():
					if track_info[key] != None:
						track_info[key] =  [x.strip() for x in track_info[key].split(",")]
				
				if key == "download_1_name":
					if track_info["categories"] != None:
						if track_info["categories"] != "Track Only":
							c = c+1
							continue
					track_name = track_info["name"].partition("-")[0].strip()
					track_name_extention = track_info["name"].partition("+",)[2].rstrip(string.digits).rstrip("$").strip()
					if track_name_extention != "":
						track_info[key] = track_name +" - " +track_name_extention  + ".zip"
					else:
						track_info[key] = track_name + ".zip"
				# if key == "mins_value(s)":
				# 	for track in os.listdir(src_track_directory):
				# 		if "-" not in track and track_info["name"] in track and track[-3:] == "wav":
				# 		# try:
		
				# 			with contextlib.closing(wave.open(src_track_directory+"/"+track,'r')) as f:
				# 				frames = f.getnframes()
				# 				rate = f.getframerate()
				# 				duration = frames / float(rate)
				# 				mins = duration//60
				# 				seconds =  (duration % 60)
				# 				track_info["mins_value(s)"] = "%02d:%02d" % (mins, seconds)
				# 		# except:
				# 				print(track + ": cant get track length possible unknown format")
				if row[c].value != None:
					if key == "genre_value(s)":
						track_info["genre_value(s)"] = row[c].value.split(", ")
					if key == "mood_value(s)":
						track_info["mood_value(s)"] = row[c].value.split(", ")
					if key == "instrument_value(s)":
						track_info["instrument_value(s)"] = row[c].value.split(", ")	
					if key == "tags":
						track_info["tags"] = row[c].value.split(", ")	
					if key == "purpose_value(s)":
						track_info["purpose_value(s)"] = row[c].value.split(", ")				
				if row[c].value == None:
					track_info[key] = ""

			c = c+1

							# print(src_track_directory+"/"+track)
							# audio = MP3(src_track_directory+"/"+track)
							# mins = audio.info.length//60
							# seconds =  (audio.info.length % 60)
							# # track_info["mins_value(s)"] = "%02d:%02d" % (mins, seconds)
							# track_info["mins_value(s)"] = audio.info.length
						



		if is_product_or_variation == "product":
			current_product_name = track_info["name"]
			categories = track_info["categories"]

		row_list.append(row[0].value)
		row_list.append(track_info)

		track_list.append(row_list)

	track_list.pop(0)

	product_variation_list = []
	all_products_variation_list=[]
	c=0

	#group a product with is variations
	for track in track_list:
		if track[0] == None:
			all_products_variation_list.append(product_variation_list)
			break

		if track[0].lower() == "product":
			c = c+1

		if c == 2:
			all_products_variation_list.append(product_variation_list)
			product_variation_list = []
			c=1
		
		product_variation_list.append(track)


	return all_products_variation_list

# pprint.pprint(read_excel("track_information_to_add_final.xlsx", "all_tracks"))
# test = read_excel("track_information.xlsx", "all_tracks")
